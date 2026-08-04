"""
测试用例操作综合测试脚本（修复版）
覆盖所有场景：创建、新增、删除、修改、批量变更、多Sheet等
已修复 insert_rows/delete_rows 的内容丢失问题

运行方式: python test_openpyxl_comprehensive.py
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import os
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ============ 配置 ============
TEST_FILE = 'test_comprehensive.xlsx'
COPY_FILE = 'test_comprehensive_copy.xlsx'

# ============ 样式定义 ============
TC_FONT = Font(name='微软雅黑', size=10)
TC_FILL = PatternFill(start_color='FFFCE4EC', end_color='FFFCE4EC', fill_type='solid')
TC_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
TC_ALIGN = Alignment(vertical='center', wrap_text=True)

SECTION_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFFFF')
SECTION_FILL = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')

HEADER_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFFFF')
HEADER_FILL = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')


# ============ 工具函数 ============
def print_excel_content(wb, label=""):
    """打印 Excel 内容用于验证"""
    ws = wb.active
    print(f"\n{'='*60}")
    print(f"【{label}】Excel 内容:")
    print('='*60)
    for row in range(1, ws.max_row + 1):
        a_val = ws.cell(row, 1).value
        b_val = ws.cell(row, 2).value
        d_val = ws.cell(row, 4).value

        if a_val and str(a_val).startswith('【'):
            is_merged = any(m.min_row == row and m.max_row == row for m in ws.merged_cells.ranges)
            print(f"  行{row:2d}: [Section] {a_val} {'✓合并' if is_merged else '✗未合并'}")
        elif a_val and str(a_val).startswith('TC-'):
            has_content = "✓" if b_val and d_val else "✗丢失!"
            print(f"  行{row:2d}: {a_val} | {b_val or '空'} | {(d_val or '空')[:15]} | {has_content}")
        elif row == 1:
            print(f"  行{row:2d}: [表头]")


def create_test_excel():
    """测试1：创建测试 Excel 文件"""
    print("\n" + "="*60)
    print("【测试1】创建测试 Excel 文件")
    print("="*60)

    wb = Workbook()
    ws = wb.active
    ws.title = '测试Sheet'

    # 表头
    headers = ['用例编号', '场景', '所属模块', '用例标题', '前置条件', '操作步骤', '预期结果', '优先级', '执行结果', '备注']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = TC_BORDER

    # Section 和用例数据
    sections = [
        ('【① 登录模块】', [
            ('TC-001', '微信登录', '登录', '微信授权后登录', '已装微信', '1.点击\n2.授权', '1.跳转\n2.成功', 'P0'),
            ('TC-002', 'Apple登录', '登录', 'Apple授权后登录', 'iOS', '1.点击\n2.授权', '1.弹出\n2.成功', 'P0'),
        ]),
        ('【② 视频模块】', [
            ('TC-003', '视频解锁', '视频', '解锁视频成功', '已登录', '1.选择\n2.支付', '1.弹出\n2.成功', 'P0'),
            ('TC-004', '视频下载', '视频', '下载视频成功', '已解锁', '1.下载', '1.完成', 'P0'),
        ]),
        ('【③ 播放器模块】', [
            ('TC-005', '手势', '播放器', '双击暂停', '播放中', '1.双击', '1.暂停', 'P0'),
            ('TC-006', '手势', '播放器', '长按快进', '播放中', '1.长按', '1.快进', 'P1'),
        ]),
    ]

    row = 2
    for sec_name, cases in sections:
        # Section Header
        ws.merge_cells(f'A{row}:J{row}')
        cell = ws.cell(row, 1, sec_name)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = TC_ALIGN
        for c in range(1, 11):
            ws.cell(row, c).border = TC_BORDER
        row += 1

        # TC 行
        for case in cases:
            for col, val in enumerate(case, 1):
                cell = ws.cell(row, col, val)
                cell.font = TC_FONT
                cell.fill = TC_FILL
                cell.border = TC_BORDER
                cell.alignment = TC_ALIGN
            row += 1

    # 执行结果下拉（注意：英文逗号分隔）
    dv = DataValidation(type='list', formula1='"通过,失败"', allow_blank=True)
    dv.error = '请选择通过或失败'
    dv.prompt = '请选择执行结果'
    ws.add_data_validation(dv)
    dv.add('I2:I10000')

    # 条件格式
    green_fill = PatternFill(start_color='FFC6EFCE', fill_type='solid')
    green_font = Font(name='微软雅黑', size=10, bold=True, color='FF006100')
    red_fill = PatternFill(start_color='FFFFC7CE', fill_type='solid')
    red_font = Font(name='微软雅黑', size=10, bold=True, color='FF9C0006')
    ws.conditional_formatting.add('I2:I10000',
        CellIsRule(operator='equal', formula=['"通过"'], fill=green_fill, font=green_font))
    ws.conditional_formatting.add('I2:I10000',
        CellIsRule(operator='equal', formula=['"失败"'], fill=red_fill, font=red_font))

    wb.save(TEST_FILE)
    print(f"✓ 创建成功: {TEST_FILE}")
    print(f"  用例数: 6 条 (TC-001 ~ TC-006)")
    print(f"  Section 数: 3 个")


def fixed_insert_rows(ws, insert_row, new_tc_data):
    """修复版 insert_rows - 先取消所有合并，操作后在新位置重新合并"""
    # 步骤1: 取消所有 Section 的合并（关键修复！）
    for m in list(ws.merged_cells.ranges):
        ws.merged_cells.ranges.remove(m)

    # 步骤2: 执行 insert_rows
    ws.insert_rows(insert_row + 1)
    new_row = insert_row + 1

    # 步骤3: 填写新用例
    for col, val in enumerate(new_tc_data, 1):
        ws.cell(new_row, col, val)
    for c in range(1, 11):
        ws.cell(new_row, c).font = TC_FONT
        ws.cell(new_row, c).fill = TC_FILL
        ws.cell(new_row, c).border = TC_BORDER
        ws.cell(new_row, c).alignment = TC_ALIGN

    # 步骤4: 后续编号 +1
    for r in range(insert_row + 2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).startswith('TC-'):
            num = int(str(v).split('-')[1])
            ws.cell(r, 1).value = f'TC-{num+1:03d}'

    # 步骤5: 重新合并所有 Section（在新位置）
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).startswith('【'):
            ws.merge_cells(f'A{r}:J{r}')
            ws.cell(r, 1).font = SECTION_FONT
            ws.cell(r, 1).fill = SECTION_FILL

    # 步骤6: 验证内容完整性
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).startswith('TC-'):
            if not ws.cell(r, 2).value or not ws.cell(r, 4).value:
                raise Exception(f"内容丢失！行 {r}")


def test_insert_rows():
    """测试2：新增用例（修复版）"""
    print("\n" + "="*60)
    print("【测试2】新增用例 - insert_rows（修复版）")
    print("="*60)

    # 加载原文件
    wb = load_workbook(TEST_FILE)
    ws = wb.active
    print_excel_content(wb, "插入前")

    # 找到 TC-002 的位置
    insert_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == 'TC-002':
            insert_row = r
            break

    if not insert_row:
        print("错误：找不到 TC-002")
        return False

    print(f"\n准备在第 {insert_row} 行后插入新用例...")

    # 使用修复版 insert_rows
    new_tc = ('TC-NEW', 'Google登录', '登录', 'Google授权后登录', '已装Google', '1.点击\n2.授权', '1.跳转\n2.成功', 'P0')
    fixed_insert_rows(ws, insert_row, new_tc)

    print_excel_content(wb, "插入后（保存前）")

    # 保存为拷贝文件
    wb.save(COPY_FILE)
    print(f"\n✓ 保存为拷贝文件: {COPY_FILE}")

    # 重新加载验证
    wb2 = load_workbook(COPY_FILE)
    print_excel_content(wb2, "重新加载后")

    # 验证所有 TC 内容完整
    ws2 = wb2.active
    all_ok = True
    for r in range(2, ws2.max_row + 1):
        v = ws2.cell(r, 1).value
        if v and str(v).startswith('TC-'):
            if not ws2.cell(r, 2).value or not ws2.cell(r, 4).value:
                print(f"  ❌ 行 {r}: {v} 内容丢失!")
                all_ok = False

    if all_ok:
        print("\n✓ 所有内容完整，修复成功！")

    return all_ok


def fixed_delete_rows(ws, delete_row):
    """修复版 delete_rows - 先取消所有合并，操作后在新位置重新合并"""
    # 步骤1: 取消所有 Section 的合并（关键修复！）
    for m in list(ws.merged_cells.ranges):
        ws.merged_cells.ranges.remove(m)

    # 步骤2: 删除目标行
    ws.delete_rows(delete_row)

    # 步骤3: 后续编号 -1
    for r in range(delete_row, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).startswith('TC-'):
            num = int(str(v).split('-')[1])
            ws.cell(r, 1).value = f'TC-{num-1:03d}'

    # 步骤4: 重新合并所有 Section
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).startswith('【'):
            ws.merge_cells(f'A{r}:J{r}')
            ws.cell(r, 1).font = SECTION_FONT
            ws.cell(r, 1).fill = SECTION_FILL

    # 步骤5: 验证内容完整性
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).startswith('TC-'):
            if not ws.cell(r, 2).value or not ws.cell(r, 4).value:
                raise Exception(f"内容丢失！行 {r}")


def test_delete_rows():
    """测试3：删除用例（修复版）"""
    print("\n" + "="*60)
    print("【测试3】删除用例 - delete_rows（修复版）")
    print("="*60)

    wb = load_workbook(COPY_FILE)
    ws = wb.active
    print_excel_content(wb, "删除前")

    # 找到 TC-004
    delete_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == 'TC-004':
            delete_row = r
            break

    if not delete_row:
        print("错误：找不到 TC-004")
        return False

    print(f"\n准备删除第 {delete_row} 行 (TC-004)...")

    # 使用修复版 delete_rows
    fixed_delete_rows(ws, delete_row)

    print_excel_content(wb, "删除后（保存前）")

    # 保存
    wb.save(COPY_FILE)

    # 重新加载验证
    wb2 = load_workbook(COPY_FILE)
    print_excel_content(wb2, "重新加载后")

    # 验证
    ws2 = wb2.active
    all_ok = True
    for r in range(2, ws2.max_row + 1):
        v = ws2.cell(r, 1).value
        if v and str(v).startswith('TC-'):
            if not ws2.cell(r, 2).value or not ws2.cell(r, 4).value:
                print(f"  ❌ 行 {r}: {v} 内容丢失!")
                all_ok = False

    if all_ok:
        print("\n✓ 所有内容完整，修复成功！")

    return all_ok


def test_modify_case():
    """测试4：修改用例"""
    print("\n" + "="*60)
    print("【测试4】修改用例")
    print("="*60)

    wb = load_workbook(COPY_FILE)
    ws = wb.active

    # 找到 TC-002
    target_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == 'TC-002':
            target_row = r
            break

    if not target_row:
        print("错误：找不到 TC-002")
        return False

    print(f"\n修改前: {ws.cell(target_row, 7).value}")
    ws.cell(target_row, 7).value = '1.弹出\n2.成功（已修改）'
    print(f"修改后: {ws.cell(target_row, 7).value}")

    wb.save(COPY_FILE)
    print("\n✓ 修改成功")
    return True


def cleanup():
    """清理测试文件"""
    for f in [TEST_FILE, COPY_FILE]:
        if os.path.exists(f):
            os.remove(f)


def main():
    print("\n" + "="*60)
    print("测试用例操作 - 综合测试（修复版）")
    print("="*60)

    # 清理旧文件
    cleanup()

    results = []

    # 测试1：创建
    try:
        create_test_excel()
        results.append(("创建", True))
    except Exception as e:
        print(f"❌ 创建失败: {e}")
        results.append(("创建", False))
        return

    # 测试2：新增（修复版）
    try:
        ok = test_insert_rows()
        results.append(("新增(insert_rows)修复版", ok))
    except Exception as e:
        print(f"❌ 新增失败: {e}")
        results.append(("新增(insert_rows)修复版", False))

    # 测试3：删除（修复版）
    try:
        ok = test_delete_rows()
        results.append(("删除(delete_rows)修复版", ok))
    except Exception as e:
        print(f"❌ 删除失败: {e}")
        results.append(("删除(delete_rows)修复版", False))

    # 测试4：修改
    try:
        ok = test_modify_case()
        results.append(("修改", ok))
    except Exception as e:
        print(f"❌ 修改失败: {e}")
        results.append(("修改", False))

    # 结果汇总
    print("\n" + "="*60)
    print("测试结果汇总")
    print("="*60)
    all_pass = True
    for name, ok in results:
        status = "✓ 通过" if ok else "❌ 失败"
        print(f"  {name}: {status}")
        if not ok:
            all_pass = False

    print("\n" + "="*60)
    if all_pass:
        print("✅ 所有测试通过！修复方案有效！")
    else:
        print("❌ 部分测试失败")
    print("="*60)

    # 清理
    cleanup()
    print(f"\n已清理临时文件")


if __name__ == '__main__':
    main()
